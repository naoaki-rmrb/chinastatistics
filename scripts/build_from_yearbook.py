"""中国統計年鑑(中文版)の .xls 群から、4テーマの省別・年次データを抽出し
Excel を生成する（NBS へ接続しない＝WAF回避のオフライン経路）。

対応表:
  gdp           <- 3-9 地区生产总值            (省別, 値＋指数 上年=100)
  retail        <- 15-12 社会消费品零售总额     (省別, 値＋比上年增长, 複数年)
  trade         <- 11-8 分地区货物进出口总额    (省別, 进出口/出口/进口)
  re_investment <- 19-15 分地区…房地产开发完成投资 (省別, 規模別合計=総投資)
  re_sold_area  <- 19-10 按用途分商品房销售面积  (全国・年次時系列)
  re_sold_value <- 19-11 按用途分商品房销售额    (全国・年次時系列)

使い方:
  python scripts/build_from_yearbook.py <年鑑zip または 中文版フォルダ> [出力xlsx]
"""

from __future__ import annotations

import io
import re
import sys
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 省の短縮名(年鑑表記) -> (中文, code, 日本語, English)
REGIONS = [
    ("全国", "000000", "全国", "China (National)"),
    ("北京", "110000", "北京", "Beijing"), ("天津", "120000", "天津", "Tianjin"),
    ("河北", "130000", "河北", "Hebei"), ("山西", "140000", "山西", "Shanxi"),
    ("内蒙古", "150000", "内モンゴル", "Inner Mongolia"),
    ("辽宁", "210000", "遼寧", "Liaoning"), ("吉林", "220000", "吉林", "Jilin"),
    ("黑龙江", "230000", "黒龍江", "Heilongjiang"),
    ("上海", "310000", "上海", "Shanghai"), ("江苏", "320000", "江蘇", "Jiangsu"),
    ("浙江", "330000", "浙江", "Zhejiang"), ("安徽", "340000", "安徽", "Anhui"),
    ("福建", "350000", "福建", "Fujian"), ("江西", "360000", "江西", "Jiangxi"),
    ("山东", "370000", "山東", "Shandong"), ("河南", "410000", "河南", "Henan"),
    ("湖北", "420000", "湖北", "Hubei"), ("湖南", "430000", "湖南", "Hunan"),
    ("广东", "440000", "広東", "Guangdong"), ("广西", "450000", "広西", "Guangxi"),
    ("海南", "460000", "海南", "Hainan"), ("重庆", "500000", "重慶", "Chongqing"),
    ("四川", "510000", "四川", "Sichuan"), ("贵州", "520000", "貴州", "Guizhou"),
    ("云南", "530000", "雲南", "Yunnan"), ("西藏", "540000", "チベット", "Tibet"),
    ("陕西", "610000", "陝西", "Shaanxi"), ("甘肃", "620000", "甘粛", "Gansu"),
    ("青海", "630000", "青海", "Qinghai"), ("宁夏", "640000", "寧夏", "Ningxia"),
    ("新疆", "650000", "新疆", "Xinjiang"),
]
NAME2CODE = {r[0]: r[1] for r in REGIONS}
CODE2ZH = {r[1]: r[0] for r in REGIONS}
CODE2JA = {r[1]: r[2] for r in REGIONS}
CODE2EN = {r[1]: r[3] for r in REGIONS}
ORDER = [r[1] for r in REGIONS]


def reg_label(code: str) -> str:
    """地域名を 中文/日本語/English の3行で返す。"""
    return f"{CODE2ZH.get(code, code)}\n{CODE2JA.get(code, '')}\n{CODE2EN.get(code, '')}"


# 系列名(指標名) の対訳: 中文 -> (日本語, English)
SERIES_I18N = {
    "地区生产总值": ("域内総生産(GRP)", "Gross Regional Product"),
    "人均地区生产总值(元)": ("一人当たりGRP(元)", "Per-capita GRP (yuan)"),
    "第一产业": ("第一次産業", "Primary industry"),
    "第二产业": ("第二次産業", "Secondary industry"),
    "第三产业": ("第三次産業", "Tertiary industry"),
    "农林牧渔业": ("農林牧漁業", "Agri./Forestry/Husbandry/Fishery"),
    "工业": ("工業", "Industry"),
    "建筑业": ("建設業", "Construction"),
    "批发和零售业": ("卸売・小売業", "Wholesale & Retail"),
    "交通运输仓储和邮政业": ("運輸・倉庫・郵政業", "Transport/Storage/Post"),
    "住宿和餐饮业": ("宿泊・飲食業", "Hotels & Catering"),
    "金融业": ("金融業", "Finance"),
    "房地产业": ("不動産業", "Real Estate"),
    "其他": ("その他", "Others"),
    "社会消费品零售总额": ("社会消費財小売総額", "Total Retail Sales of Consumer Goods"),
    "货物进出口": ("貨物輸出入", "Goods Imports & Exports"),
    "货物出口": ("貨物輸出", "Goods Exports"),
    "货物进口": ("貨物輸入", "Goods Imports"),
    "房地产开发完成投资": ("不動産開発完成投資", "Real Estate Development Investment"),
    "商品房销售面积": ("商品住宅販売面積", "Floor Space of Buildings Sold"),
    "商品房销售额": ("商品住宅販売額", "Sales of Buildings (value)"),
    "全社会固定资产投资(亿元)": ("全社会固定資産投資(億元)", "Total Fixed Asset Investment (100M yuan)"),
    "全社会FAI 比上年增长(%)": ("全社会FAI 前年比(%)", "Total FAI YoY (%)"),
    "房地产开发投资(亿元)": ("不動産開発投資(億元)", "Real Estate Dev. Investment (100M yuan)"),
    "房地产开发投资 比上年增长(%)": ("不動産開発投資 前年比(%)", "RE Dev. Investment YoY (%)"),
    "全部投资(亿元)": ("全体投資(億元)", "Total Investment (100M yuan)"),
    "第一产业(亿元)": ("第一次産業(億元)", "Primary industry (100M yuan)"),
    "第二产业(亿元)": ("第二次産業(億元)", "Secondary industry (100M yuan)"),
    "第三产业(亿元)": ("第三次産業(億元)", "Tertiary industry (100M yuan)"),
}


def ser_label(zh: str) -> str:
    """系列名を 中文/日本語/English の3行で返す（未登録は中文のみ）。"""
    ja, en = SERIES_I18N.get(zh, ("", ""))
    parts = [zh]
    if ja:
        parts.append(ja)
    if en:
        parts.append(en)
    return "\n".join(parts)


def norm_region(s) -> str | None:
    """セル値を省コードに正規化。省名でなければ None。"""
    if not isinstance(s, str):
        return None
    key = re.sub(r"\s", "", s)
    key = key.replace("　", "")
    # 「北  京」→「北京」など
    if key in NAME2CODE:
        return NAME2CODE[key]
    # 一部表記ゆれ（黑龙江/黑龍江 等）は先頭2文字で近似
    for name, code in NAME2CODE.items():
        if key and (key.startswith(name) or name.startswith(key)):
            return code
    return None


def _num(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def load_zip(path: Path) -> dict[str, bytes]:
    """zip/フォルダから {デコード済みパス: bytes} を返す（中文版のみ）。"""
    out: dict[str, bytes] = {}
    if path.is_dir():
        for p in path.rglob("*.xls"):
            out[str(p)] = p.read_bytes()
        return out
    z = zipfile.ZipFile(path)
    for n in z.namelist():
        if not n.lower().endswith(".xls"):
            continue
        try:
            d = n.encode("cp437").decode("utf-8")
        except Exception:
            try:
                d = n.encode("cp437").decode("gbk")
            except Exception:
                d = n
        if "中文版" in d or path.is_dir():
            out[d] = z.read(n)
    return out


def find(files: dict[str, bytes], fragment: str) -> bytes | None:
    for d, b in files.items():
        if fragment in d:
            return b
    return None


def read_xls(b: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(b), header=None, engine="xlrd")


# ---------------------------------------------------------------------
# 各表の抽出（tidy レコード: theme, region_code, year, series, value, yoy）
# ---------------------------------------------------------------------
def parse_retail(b: bytes) -> list[dict]:
    df = read_xls(b)
    # 年は row2 の 1,3 列に入る（例 2021, 2022）
    years = {}
    for c in range(1, df.shape[1]):
        y = _num(df.iat[2, c]) if df.shape[0] > 2 else None
        if y and 1990 <= y <= 2100:
            years[c] = int(y)
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        for c, yr in years.items():
            val = _num(df.iat[i, c])
            yoy = _num(df.iat[i, c + 1]) if c + 1 < df.shape[1] else None
            if val is not None:
                recs.append(dict(theme="retail", region_code=code, year=yr,
                                 series="社会消费品零售总额", value=val, yoy=yoy))
    return recs


def parse_gdp(b: bytes) -> list[dict]:
    df = read_xls(b)
    year = 2022
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        val = _num(df.iat[i, 1])  # 地区生产总值(亿元)
        # 指数(上年=100) の「地区生产总值」列。3-9 標準様式では col18。
        # 指数ブロック(地区GDP/一産/二産/三産/人均)の先頭 = 地区生产总值指数。
        idx = None
        if df.shape[1] > 18:
            v = _num(df.iat[i, 18])
            if v is not None and 80 <= v <= 130:
                idx = v
        if idx is None:
            # フォールバック: 指数ブロックらしき最初の 80〜130 を左から探す
            # (構成比の 80〜130 誤検出を避けるため col15 以降に限定)
            for c in range(15, df.shape[1]):
                v = _num(df.iat[i, c])
                if v is not None and 95 <= v <= 130:
                    idx = v
                    break
        yoy = round(idx - 100, 1) if idx is not None else None
        if val is not None:
            recs.append(dict(theme="gdp", region_code=code, year=year,
                             series="地区生产总值", value=val, yoy=yoy))
    return recs


def parse_trade(b: bytes) -> list[dict]:
    df = read_xls(b)
    year = 2022
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        for c, name in [(1, "货物进出口"), (2, "货物出口"), (3, "货物进口")]:
            val = _num(df.iat[i, c])
            if val is not None:
                recs.append(dict(theme="trade", region_code=code, year=year,
                                 series=name, value=val, yoy=None))
    return recs


def parse_re_investment(b: bytes) -> list[dict]:
    df = read_xls(b)
    year = 2022
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        # 規模別列(1..)の合計 = 房地产开发完成投资 総額
        vals = [_num(df.iat[i, c]) for c in range(1, df.shape[1])]
        vals = [v for v in vals if v is not None]
        if vals:
            recs.append(dict(theme="re_investment", region_code=code, year=year,
                             series="房地产开发完成投资", value=sum(vals), yoy=None))
    return recs


def parse_national_timeseries(b: bytes, theme: str, series: str) -> list[dict]:
    """19-10/19-11: col0=年, col1=総額（全国時系列）。"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        y = _num(df.iat[i, 0])
        if y is None or not (1990 <= y <= 2100):
            continue
        val = _num(df.iat[i, 1])
        if val is not None:
            recs.append(dict(theme=theme, region_code="000000", year=int(y),
                             series=series, value=val, yoy=None))
    return recs


def parse_percapita_gdp(b: bytes) -> list[dict]:
    """3-9 の 人均地区生产总值(元) 列(col13) を省別に抽出。"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        v = _num(df.iat[i, 14]) if df.shape[1] > 14 else None  # (元) 人均列
        if v is not None and v > 1000:  # 人均は元単位で大きい値
            recs.append(dict(theme="gdp_percapita", region_code=code, year=2022,
                             series="人均地区生产总值(元)", value=v, yoy=None))
    return recs


def parse_overview(b: bytes, theme: str, value_col: int) -> list[dict]:
    """全国推移表: col0=年(範囲文字列可), value_col=金額。戻り: {theme,year_label,value}"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        y = df.iat[i, 0]
        if not isinstance(y, (str, int, float)):
            continue
        ylabel = str(y).strip()
        if not re.match(r"^\d{4}", ylabel):
            continue
        val = _num(df.iat[i, value_col])
        if val is not None:
            recs.append(dict(theme=theme, year_label=ylabel, value=val))
    return recs


COUNTRY_SKIP = {"国别(地区)", "国家(地区)", "单位:万美元", "单位：万美元"}


def parse_country(b: bytes, dataset: str, year_cols: dict) -> list[dict]:
    """国・地域別表: col0=国名(#や大陸小計含む), year_cols={列:ラベル}。"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        name = df.iat[i, 0]
        if not isinstance(name, str):
            continue
        cn = re.sub(r"\s", "", name).lstrip("#").strip()
        if not cn or cn in COUNTRY_SKIP:
            continue
        # 表タイトル行(例「11-14按国别…」)や見出しを除外
        if re.match(r"^\d", cn) or "直接投资" in cn or "单位" in cn:
            continue
        for col, label in year_cols.items():
            if col >= df.shape[1]:
                continue
            val = _num(df.iat[i, col])
            if val is not None:
                recs.append(dict(dataset=dataset, country=cn, period=label, value=val))
    return recs


GDP_IND_COLS = {
    2: "第一产业", 3: "第二产业", 4: "第三产业",
    5: "农林牧渔业", 6: "工业", 7: "建筑业", 8: "批发和零售业",
    9: "交通运输仓储和邮政业", 10: "住宿和餐饮业", 11: "金融业",
    12: "房地产业", 13: "其他",
}


def parse_gdp_by_industry(b: bytes) -> list[dict]:
    """3-9 から省別の産業別付加価値(億元, 2022)を抽出。"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        code = norm_region(df.iat[i, 0])
        if not code:
            continue
        for col, name in GDP_IND_COLS.items():
            if col >= df.shape[1]:
                continue
            v = _num(df.iat[i, col])
            if v is not None:
                recs.append(dict(theme="gdp_industry", region_code=code, year=2022,
                                 series=name, value=v, yoy=None))
    return recs


def parse_timeseries_multi(b: bytes, theme: str, colmap: dict) -> list[dict]:
    """全国推移: col0=年, colmap={列:系列名}。戻り {theme,year,series,value}"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        y = _num(df.iat[i, 0])
        if y is None or not (1978 <= y <= 2100):
            continue
        for col, name in colmap.items():
            if col >= df.shape[1]:
                continue
            v = _num(df.iat[i, col])
            if v is not None:
                recs.append(dict(theme=theme, year=int(y), series=name, value=v))
    return recs


def parse_goods(b: bytes, dataset: str) -> list[dict]:
    """11-6/11-7: col0=品名(＋単位), col1=数量, col2=金额万元, col3=金额万美元。"""
    df = read_xls(b)
    recs = []
    for i in range(df.shape[0]):
        name = df.iat[i, 0]
        if not isinstance(name, str):
            continue
        nm = re.sub(r"\s+", " ", name).strip()
        if not nm or nm.startswith("11-") or nm.replace(" ", "") in ("品名",):
            continue
        qty = _num(df.iat[i, 1])
        cny = _num(df.iat[i, 2])
        usd = _num(df.iat[i, 3]) if df.shape[1] > 3 else None
        if qty is None and cny is None and usd is None:
            continue
        recs.append(dict(dataset=dataset, item=nm, qty=qty, cny=cny, usd=usd))
    return recs


# ---------------------------------------------------------------------
# Excel 出力（省別は地区を行に、年を列に）
# ---------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build(zip_path: str, out_path: str) -> None:
    files = load_zip(Path(zip_path))
    if not files:
        raise SystemExit(f"年鑑ファイルが見つかりません: {zip_path}")

    recs: list[dict] = []
    tbl = find(files, "15-12_社会消费品零售总额")
    if tbl is not None:
        recs += parse_retail(tbl)
    tbl = find(files, "3-9_地区生产总值")
    if tbl is not None:
        recs += parse_gdp(tbl)
    tbl = find(files, "11-8_分地区货物进出口总额")
    if tbl is not None:
        recs += parse_trade(tbl)
    tbl = find(files, "19-15_分地区按项目规模分房地产开发完成投资")
    if tbl is not None:
        recs += parse_re_investment(tbl)
    tbl = find(files, "19-10_按用途分商品房销售面积")
    if tbl is not None:
        recs += parse_national_timeseries(tbl, "re_sold_area", "商品房销售面积")
    tbl = find(files, "19-11_按用途分商品房销售额")
    if tbl is not None:
        recs += parse_national_timeseries(tbl, "re_sold_value", "商品房销售额")
    # 一人当たりGDP（省別）・GDP産業別（省別）
    tbl = find(files, "3-9_地区生产总值")
    if tbl is not None:
        recs += parse_percapita_gdp(tbl)
        recs += parse_gdp_by_industry(tbl)

    # 投資の全国推移（対中FDI / 対外ODI）
    overview: list[dict] = []
    tbl = find(files, "11-13_外商直接投资情况")
    if tbl is not None:
        overview += parse_overview(tbl, "fdi_in", value_col=2)   # 实际使用外资(亿美元)
    tbl = find(files, "11-18_对外直接投资")
    if tbl is not None:
        overview += parse_overview(tbl, "odi_out", value_col=1)  # 対外投資流量(亿美元)

    # 投資の国・地域別（どこから対中FDI / どこへ対外ODI）
    country: list[dict] = []
    tbl = find(files, "11-14_按国别")
    if tbl is not None:
        country += parse_country(tbl, "対中投資FDI(国別)", {1: "2021", 2: "2022"})
    tbl = find(files, "11-19_按主要国别")
    if tbl is not None:
        country += parse_country(tbl, "対外投資ODI(国別)",
                                 {1: "2020流量", 2: "2021流量", 3: "2021末存量"})

    # 固定資産投資（全国推移・産業別）と 貨物別輸出入
    ts_multi: list[dict] = []
    tbl = find(files, "10-1_全社会固定资产投资")
    if tbl is not None:
        ts_multi += parse_timeseries_multi(tbl, "固定資産投資(全国推移)", {
            1: "全社会固定资产投资(亿元)", 2: "全社会FAI 比上年增长(%)",
            3: "房地产开发投资(亿元)", 4: "房地产开发投资 比上年增长(%)"})
    tbl = find(files, "10-4_三次产业固定资产投资")
    if tbl is not None:
        ts_multi += parse_timeseries_multi(tbl, "固定資産投資(産業別・推移)", {
            1: "全部投资(亿元)", 2: "第一产业(亿元)",
            3: "第二产业(亿元)", 4: "第三产业(亿元)"})

    goods: list[dict] = []
    tbl = find(files, "11-6_出口主要货物数量和金额")
    if tbl is not None:
        goods += parse_goods(tbl, "輸出 貨物別")
    tbl = find(files, "11-7_进口主要货物数量和金额")
    if tbl is not None:
        goods += parse_goods(tbl, "輸入 貨物別")

    df = pd.DataFrame.from_records(recs)
    if df.empty:
        raise SystemExit("抽出できたデータがありません。")
    df_ov = pd.DataFrame.from_records(overview)
    df_cty = pd.DataFrame.from_records(country)
    df_ts = pd.DataFrame.from_records(ts_multi)
    df_goods = pd.DataFrame.from_records(goods)

    from openpyxl import Workbook
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "説明"
    info = [
        ("中国 省別・年次データ（統計年鑑2023より抽出）", TITLE_FONT),
        ("China provincial annual data extracted from Statistical Yearbook 2023", None),
        ("", None),
        ("出所: 中国統計年鑑2023（中文版）公式.xls", None),
        ("※ 年鑑は年次。省別は概ね2022年（小売は2021・2022）。", None),
        ("※ 前年比(YoY)は年鑑の公式値: 小売=比上年增长, GDP=指数(上年=100)より。", None),
        ("※ 月次データは別途 NBS(easyquery) から取得（VPN/プロキシ経由）。", None),
    ]
    for i, (t, f) in enumerate(info, 1):
        c = ws0.cell(i, 1, t)
        if f:
            c.font = f
    ws0.column_dimensions["A"].width = 80

    # テーマごとにシート（地区×年 の 値、YoY があれば併記）
    theme_titles = {
        "gdp": "GDP 地区生产总值", "gdp_percapita": "一人当たりGDP 人均",
        "gdp_industry": "GDP産業別(省別)",
        "retail": "小売 社会消费品零售总额",
        "trade": "貿易 货物进出口", "re_investment": "不動産 开发完成投资",
        "re_sold_area": "全国 商品房销售面积", "re_sold_value": "全国 商品房销售额",
    }
    for theme, title in theme_titles.items():
        sub = df[df["theme"] == theme]
        if sub.empty:
            continue
        ws = wb.create_sheet(title[:31])
        _write_theme(ws, title, sub)

    # 投資の全国推移（年×金額）
    if not df_ov.empty:
        ov_titles = {"fdi_in": "対中投資FDI(全国推移)", "odi_out": "対外投資ODI(全国推移)"}
        for theme, title in ov_titles.items():
            sub = df_ov[df_ov["theme"] == theme]
            if sub.empty:
                continue
            ws = wb.create_sheet(title[:31])
            _write_overview(ws, title, sub)

    # 投資の国・地域別（国×期間）
    if not df_cty.empty:
        for dataset in df_cty["dataset"].unique():
            sub = df_cty[df_cty["dataset"] == dataset]
            ws = wb.create_sheet(str(dataset)[:31])
            _write_country(ws, str(dataset), sub)

    # 固定資産投資（全国推移・産業別）: 年×系列
    if not df_ts.empty:
        for theme in df_ts["theme"].unique():
            sub = df_ts[df_ts["theme"] == theme]
            ws = wb.create_sheet(str(theme)[:31])
            _write_timeseries(ws, str(theme), sub)

    # 貨物別輸出入: 品目×[数量/金額]
    if not df_goods.empty:
        for dataset in df_goods["dataset"].unique():
            sub = df_goods[df_goods["dataset"] == dataset]
            ws = wb.create_sheet(str(dataset)[:31])
            _write_goods(ws, str(dataset), sub)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"生成: {out_path}  （省別{len(df)}, 推移{len(df_ov)}, 国別{len(df_cty)}, "
          f"FAI{len(df_ts)}, 貨物{len(df_goods)} レコード）")


def _write_timeseries(ws, title, sub) -> None:
    """全国推移(多系列): 年を行、系列を列。"""
    ws.cell(1, 1, title).font = TITLE_FONT
    series = list(dict.fromkeys(sub["series"].tolist()))
    ws.cell(3, 1, "年 / Year").font = HDR_FONT
    ws.cell(3, 1).fill = HDR_FILL
    for j, s in enumerate(series):
        cell = ws.cell(3, 2 + j, ser_label(s))
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    years = sorted(sub["year"].unique().tolist())
    r = 4
    for y in years:
        ws.cell(r, 1, y)
        for j, s in enumerate(series):
            rows = sub[(sub.year == y) & (sub.series == s)]
            if not rows.empty:
                cell = ws.cell(r, 2 + j, float(rows.iloc[0]["value"]))
                cell.number_format = '0.0"%"' if "增长" in s or "%" in s else "#,##0.0"
        r += 1
    ws.freeze_panes = "B4"
    ws.row_dimensions[3].height = 66
    ws.column_dimensions["A"].width = 8
    for j in range(len(series)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 24


def _write_goods(ws, title, sub) -> None:
    """貨物別: 品目を行、[数量/金額(万元)/金額(万美元)]。"""
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, "数量の単位は品目名に付記。金額は万元人民币/万美元。").font = \
        Font(size=9, color="595959")
    heads = ["品名 / Item", "数量 Qty", "金額(万元)", "金額(万美元)"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(4, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    r = 5
    for row in sub.itertuples(index=False):
        ws.cell(r, 1, row.item)
        if pd.notna(row.qty):
            ws.cell(r, 2, float(row.qty)).number_format = "#,##0"
        if pd.notna(row.cny):
            ws.cell(r, 3, float(row.cny)).number_format = "#,##0"
        if pd.notna(row.usd):
            ws.cell(r, 4, float(row.usd)).number_format = "#,##0"
        r += 1
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 26
    for c in "BCD":
        ws.column_dimensions[c].width = 14


def _write_overview(ws, title, sub) -> None:
    """全国推移: 年ラベル×金額(亿美元)。"""
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, "単位: 億米ドル / 100 mil. USD").font = Font(size=9, color="595959")
    ws.cell(4, 1, "年 / Year").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    ws.cell(4, 2, "金額").font = HDR_FONT
    ws.cell(4, 2).fill = HDR_FILL
    r = 5
    for row in sub.itertuples(index=False):
        ws.cell(r, 1, row.year_label)
        ws.cell(r, 2, float(row.value)).number_format = "#,##0.0"
        r += 1
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14


def _write_country(ws, title, sub) -> None:
    """国・地域別: 国を行、期間を列。"""
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, "単位: 万米ドル / 10,000 USD").font = Font(size=9, color="595959")
    periods = list(dict.fromkeys(sub["period"].tolist()))
    ws.cell(4, 1, "国・地域 / Country").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    for j, p in enumerate(periods):
        cell = ws.cell(4, 2 + j, p)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    countries = list(dict.fromkeys(sub["country"].tolist()))
    r = 5
    for cn in countries:
        ws.cell(r, 1, cn)
        for j, p in enumerate(periods):
            rows = sub[(sub.country == cn) & (sub.period == p)]
            if not rows.empty:
                ws.cell(r, 2 + j, float(rows.iloc[0]["value"])).number_format = "#,##0"
        r += 1
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 18
    for j in range(len(periods)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 14


def _write_theme(ws, title, sub) -> None:
    ws.cell(1, 1, title).font = TITLE_FONT
    series_list = list(dict.fromkeys(sub["series"].tolist()))
    years = sorted(sub["year"].unique().tolist())
    has_yoy = sub["yoy"].notna().any()

    # ヘッダ: 地区 | (系列×年 の 値) ... | (YoY×年) ...
    row = 3
    hc = ws.cell(row, 1, "地区\n地域\nRegion")
    hc.font = HDR_FONT; hc.fill = HDR_FILL; hc.alignment = CENTER
    col = 2
    colmap = {}
    for s in series_list:
        for y in years:
            cell = ws.cell(row, col, f"{ser_label(s)}\n{y} 値")
            cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER
            colmap[("val", s, y)] = col
            col += 1
    if has_yoy:
        for s in series_list:
            for y in years:
                if sub[(sub.series == s) & (sub.year == y)]["yoy"].notna().any():
                    cell = ws.cell(row, col, f"{ser_label(s)}\n{y} 前年比%")
                    cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER
                    colmap[("yoy", s, y)] = col
                    col += 1

    # 地区は年鑑順（中文/日本語/English を1セルに3行）
    present = set(sub["region_code"])
    r = row + 1
    for code in ORDER:
        if code not in present:
            continue
        rc = ws.cell(r, 1, reg_label(code))
        rc.alignment = Alignment(vertical="center", wrap_text=True)
        for (kind, s, y), c in colmap.items():
            rows = sub[(sub.region_code == code) & (sub.series == s) & (sub.year == y)]
            if rows.empty:
                continue
            v = rows.iloc[0]["yoy" if kind == "yoy" else "value"]
            if pd.notna(v):
                cell = ws.cell(r, c, float(v))
                cell.number_format = '0.0"%"' if kind == "yoy" else "#,##0.0"
        r += 1

    ws.freeze_panes = "B4"
    ws.row_dimensions[row].height = 66   # 3言語＋年でヘッダを高く
    ws.column_dimensions["A"].width = 16
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 16


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    zp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "output/china_yearbook_provincial.xlsx"
    build(zp, out)
