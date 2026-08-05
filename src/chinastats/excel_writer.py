"""DataFrame から前年比/前月比付きの Excel を生成する。

シート構成:
  説明          : データ源・注意・最終更新・解決済み指標コード
  指標一覧      : 3言語(中/日/英)の指標名と単位の対応表
  改定検知      : 前回スナップショットからの過去値改定（下方改定の炙り出し）
  <各指標>      : 地区(列)×時点(行) の行列を、指標ごとに縦に積む
                  値 / 公式前年比 / 計算前年比 / 乖離 / 前月比
                  （累計指標は単月・単月前月比も）
                  ※ 計算前年比・乖離・前月比は Excel の「数式」で入れる
                    → 表の上で去年値と今年値を突き合わせて計算するため、
                      改定の有無を自分の目で検証できる。
"""

from __future__ import annotations

import logging
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---- スタイル ----
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SECT_FILL = PatternFill("solid", fgColor="D9E1F2")
SECT_FONT = Font(bold=True, size=10, color="1F4E78")
TITLE_FONT = Font(bold=True, size=12)
SUB_FONT = Font(size=9, color="595959")
REGION_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

VALUE_FMT = "#,##0.0"
PCT_FMT = '0.0"%"'

# 各指標シートに積む行列ブロックの定義
# (見出し, どの列から, 数値フォーマット, 数式か)
BASE_BLOCKS = [
    ("値 / Value", "value", VALUE_FMT, False),
    ("公式前年比 / Official YoY", "official_yoy_pct", PCT_FMT, False),
    ("計算前年比 / Computed YoY", "computed_yoy", PCT_FMT, True),
    ("乖離(計算-公式) / Gap", "yoy_gap", PCT_FMT, True),
    ("前月比 / MoM", "mom", PCT_FMT, True),
]
CUMULATIVE_EXTRA = [
    ("単月(復元) / Single-month", "single_month", VALUE_FMT, False),
    ("単月前月比 / Single-month MoM", "single_mom", PCT_FMT, False),
]


def _region_order(regions_config: dict[str, Any]) -> list[dict[str, str]]:
    order = []
    nat = regions_config.get("national")
    if nat:
        order.append({"code": nat["code"], "zh": nat["name_zh"], "ja": nat.get("name_ja", nat["name_zh"])})
    for r in regions_config.get("provinces", []):
        order.append({"code": r["code"], "zh": r["name_zh"], "ja": r["name_ja"]})
    return order


def _write_matrix(
    ws: Worksheet,
    start_row: int,
    title: str,
    metric: str,
    periods: list[str],
    regions: list[dict[str, str]],
    data: dict[tuple[str, str], Any],
    number_fmt: str,
    value_cell_index: dict[tuple[str, str], str] | None,
    prev_year_period: dict[str, str | None],
    prev_period: dict[str, str | None],
    official_cell_index: dict[tuple[str, str], str] | None,
    computed_cell_index: dict[tuple[str, str], str] | None,
    as_formula: bool,
) -> tuple[int, dict[tuple[str, str], str]]:
    """1つの行列ブロックを書き、次の開始行と (period,region)->セル索引を返す。

    data: {(period, region_code): 値}  ← as_formula=False のときの直接値
    value_cell_index: 値ブロックのセル参照 {(period,region): "B5"} 等
                      計算前年比/前月比の数式が参照する。
    """
    n_region = len(regions)
    r = start_row

    # セクション見出し
    ws.cell(r, 1, title).font = SECT_FONT
    for c in range(1, n_region + 2):
        ws.cell(r, c).fill = SECT_FILL
    r += 1

    # ヘッダ行（地区: 日本語 / 中文 の2段）
    ws.cell(r, 1, "年月 / Period").font = HDR_FONT
    ws.cell(r, 1).fill = HDR_FILL
    ws.cell(r, 1).alignment = CENTER
    for j, reg in enumerate(regions):
        c = 2 + j
        cell = ws.cell(r, c, f"{reg['ja']}\n{reg['zh']}")
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    header_row = r
    r += 1

    cell_index: dict[tuple[str, str], str] = {}
    for period in periods:
        ws.cell(r, 1, period).alignment = LEFT
        ws.cell(r, 1).font = Font(size=9)
        for j, reg in enumerate(regions):
            c = 2 + j
            code = reg["code"]
            cell = ws.cell(r, c)
            cell.number_format = number_fmt
            cell.border = BORDER
            cell_ref = f"{get_column_letter(c)}{r}"
            cell_index[(period, code)] = cell_ref

            if not as_formula:
                v = data.get((period, code))
                if v is not None and pd.notna(v):
                    cell.value = float(v)
            else:
                formula = _formula_for(
                    metric, period, code, c, r,
                    value_cell_index, official_cell_index, computed_cell_index,
                    prev_year_period, prev_period,
                )
                if formula:
                    cell.value = formula
        r += 1

    ws.freeze_panes = ws.cell(header_row + 1, 2)
    return r + 1, cell_index  # 1行空ける


def _formula_for(
    metric: str,
    period: str,
    code: str,
    col: int,
    row: int,
    value_idx: dict[tuple[str, str], str] | None,
    official_idx: dict[tuple[str, str], str] | None,
    computed_idx: dict[tuple[str, str], str] | None,
    prev_year_period: dict[str, str | None],
    prev_period: dict[str, str | None],
) -> str | None:
    """計算列の Excel 数式文字列を返す。参照先が無ければ None。"""
    if metric == "computed_yoy" and value_idx is not None:
        py = prev_year_period.get(period)
        if py is None:
            return None
        cur = value_idx.get((period, code))
        prev = value_idx.get((py, code))
        if cur and prev:
            return f"=IF(OR({prev}=\"\",{prev}=0),\"\",({cur}/{prev}-1)*100)"
        return None
    if metric == "mom" and value_idx is not None:
        pp = prev_period.get(period)
        if pp is None:
            return None
        cur = value_idx.get((period, code))
        prev = value_idx.get((pp, code))
        if cur and prev:
            return f"=IF(OR({prev}=\"\",{prev}=0),\"\",({cur}/{prev}-1)*100)"
        return None
    if metric == "yoy_gap" and computed_idx is not None and official_idx is not None:
        comp = computed_idx.get((period, code))
        off = official_idx.get((period, code))
        if comp and off:
            return f"=IF(OR({comp}=\"\",{off}=\"\"),\"\",{comp}-{off})"
        return None
    return None


def _build_period_maps(periods: list[str], freq: str) -> tuple[dict, dict]:
    """period -> 前年同期 period / 前期 period のマップを作る。"""
    present = set(periods)
    prev_year: dict[str, str | None] = {}
    prev_p: dict[str, str | None] = {}
    for p in periods:
        if freq == "monthly":
            y, m = p.split("-")
            py = f"{int(y)-1}-{m}"
            prev_year[p] = py if py in present else None
            mm = int(m)
            if mm == 1:
                pp = f"{int(y)-1}-12"
            else:
                pp = f"{y}-{mm-1:02d}"
            prev_p[p] = pp if pp in present else None
        else:  # quarterly "YYYY-Qn"
            y, q = p.split("-Q")
            py = f"{int(y)-1}-Q{q}"
            prev_year[p] = py if py in present else None
            qq = int(q)
            if qq == 1:
                pp = f"{int(y)-1}-Q4"
            else:
                pp = f"{y}-Q{qq-1}"
            prev_p[p] = pp if pp in present else None
    return prev_year, prev_p


def _sheet_name(key: str, name_ja: str) -> str:
    """Excel シート名（31文字制限・禁則文字を除去）。"""
    raw = f"{key}"
    for ch in "[]:*?/\\":
        raw = raw.replace(ch, "")
    return raw[:31]


def write_workbook(
    df: pd.DataFrame,
    config: dict[str, Any],
    regions_config: dict[str, Any],
    revisions: pd.DataFrame,
    meta: dict[str, Any],
    out_path: str,
) -> None:
    """Excel ワークブックを書き出す。"""
    wb = Workbook()
    _write_readme(wb.active, meta)
    _write_indicator_list(wb.create_sheet("指標一覧"), config)
    _write_revisions(wb.create_sheet("改定検知"), revisions)

    regions = _region_order(regions_config)
    ind_defs = {i["key"]: i for i in config["indicators"]}

    for key, ind in ind_defs.items():
        sub = df[df["indicator"] == key]
        if sub.empty:
            continue
        ws = wb.create_sheet(_sheet_name(key, ind["name_ja"]))
        _write_indicator_sheet(ws, ind, sub, regions)

    wb.save(out_path)
    logger.info("Excel を書き出しました: %s", out_path)


def _write_indicator_sheet(
    ws: Worksheet,
    ind: dict[str, Any],
    sub: pd.DataFrame,
    regions: list[dict[str, str]],
) -> None:
    freq = ind.get("frequency", ind.get("freq", "monthly"))
    is_cumulative = ind.get("level_kind") == "cumulative"

    # タイトル（3言語＋単位）
    ws.cell(1, 1, ind["name_zh"]).font = TITLE_FONT
    ws.cell(2, 1, f"{ind['name_ja']}  |  {ind['name_en']}").font = SUB_FONT
    unit = f"単位: {ind.get('unit_zh','')} / {ind.get('unit_ja','')} / {ind.get('unit_en','')}"
    ws.cell(3, 1, unit).font = SUB_FONT

    periods = sorted(sub["period"].unique().tolist())
    prev_year, prev_p = _build_period_maps(periods, freq)

    # (period, region) -> 各値
    def lut(col: str) -> dict[tuple[str, str], Any]:
        d: dict[tuple[str, str], Any] = {}
        for row in sub.itertuples(index=False):
            d[(row.period, row.region_code)] = getattr(row, col)
        return d

    value_lut = lut("value")
    official_lut = lut("official_yoy_pct")
    single_lut = lut("single_month")
    single_mom_lut = lut("single_mom")

    blocks = list(BASE_BLOCKS)
    if is_cumulative:
        blocks = blocks + CUMULATIVE_EXTRA

    r = 5
    value_idx = official_idx = computed_idx = None
    for title, metric, fmt, as_formula in blocks:
        data_map: dict[tuple[str, str], Any] = {}
        if metric == "value":
            data_map = value_lut
        elif metric == "official_yoy_pct":
            data_map = official_lut
        elif metric == "single_month":
            data_map = single_lut
        elif metric == "single_mom":
            data_map = single_mom_lut

        r, idx = _write_matrix(
            ws, r, title, metric, periods, regions, data_map, fmt,
            value_cell_index=value_idx,
            prev_year_period=prev_year,
            prev_period=prev_p,
            official_cell_index=official_idx,
            computed_cell_index=computed_idx,
            as_formula=as_formula,
        )
        if metric == "value":
            value_idx = idx
        elif metric == "official_yoy_pct":
            official_idx = idx
        elif metric == "computed_yoy":
            computed_idx = idx

    # 列幅
    ws.column_dimensions["A"].width = 11
    for j in range(len(regions)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11


def _write_readme(ws: Worksheet, meta: dict[str, Any]) -> None:
    ws.title = "説明"
    lines = [
        ("中国 全国・省別 経済指標データ", TITLE_FONT),
        ("China National & Provincial Economic Indicators", SUB_FONT),
        ("", None),
        (f"最終更新 / Last update: {meta.get('updated_at','')}", None),
        (f"データ源 / Source: 国家統計局 (NBS) data.stats.gov.cn", None),
        (f"取得期間(月次) / Monthly window: {meta.get('sj_monthly','')}", None),
        (f"取得期間(四半期) / Quarterly window: {meta.get('sj_quarterly','')}", None),
        ("", None),
        ("■ 前年比の2系統 / Two YoY columns", SECT_FONT),
        ("・公式前年比 = NBS が発表する同比をそのまま取得", None),
        ("・計算前年比 = 表の上で 今年値÷去年同月値−1 を数式計算", None),
        ("・乖離 = 計算 − 公式。大きな乖離は前年値の改定を示唆", None),
        ("・改定検知シート = 前回コミット時点の過去値との差分を表示", None),
        ("", None),
        ("■ 注意 / Caveats", SECT_FONT),
        ("・GDP は四半期のみ（月次は存在しない）", None),
        ("・月次は概ね1990年代以降。1945〜80年代の月次は存在しない", None),
        ("・1月は春節の影響で単月非公表(1-2月累計)のことが多い", None),
        ("・投資・販売系は累計値。単月は累計差分で復元", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(i, 1, text)
        if font:
            cell.font = font
    ws.column_dimensions["A"].width = 70

    # 解決済みコード
    resolved = meta.get("resolved_codes") or []
    if resolved:
        base = len(lines) + 2
        ws.cell(base, 1, "解決済み指標コード / Resolved codes").font = SECT_FONT
        ws.cell(base + 1, 1, "indicator").font = HDR_FONT
        ws.cell(base + 1, 2, "db").font = HDR_FONT
        ws.cell(base + 1, 3, "role").font = HDR_FONT
        ws.cell(base + 1, 4, "zb_code").font = HDR_FONT
        ws.cell(base + 1, 5, "matched_name").font = HDR_FONT
        for k, rec in enumerate(resolved):
            row = base + 2 + k
            ws.cell(row, 1, rec.get("indicator"))
            ws.cell(row, 2, rec.get("db"))
            ws.cell(row, 3, rec.get("role"))
            ws.cell(row, 4, rec.get("code"))
            ws.cell(row, 5, rec.get("matched"))
        for col, w in zip("BCDE", (10, 10, 16, 32)):
            ws.column_dimensions[col].width = w


def _write_indicator_list(ws: Worksheet, config: dict[str, Any]) -> None:
    headers = ["key", "中文 / Chinese", "日本語 / Japanese", "English",
               "頻度/Freq", "単位(中)", "単位(日)", "単位(英)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    row = 2
    for ind in config["indicators"]:
        ws.cell(row, 1, ind["key"])
        ws.cell(row, 2, ind["name_zh"])
        ws.cell(row, 3, ind["name_ja"])
        ws.cell(row, 4, ind["name_en"])
        ws.cell(row, 5, ind.get("frequency", ind.get("freq")))
        ws.cell(row, 6, ind.get("unit_zh"))
        ws.cell(row, 7, ind.get("unit_ja"))
        ws.cell(row, 8, ind.get("unit_en"))
        row += 1
    widths = [14, 30, 30, 40, 10, 12, 12, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def _write_revisions(ws: Worksheet, revisions: pd.DataFrame) -> None:
    headers = ["指標/indicator", "名称/name", "地区/region", "期間/period",
               "旧値/old", "新値/new", "差分/diff", "変化率%/pct"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    if revisions is None or revisions.empty:
        ws.cell(2, 1, "（前回スナップショットとの差分なし / 初回実行）").font = SUB_FONT
    else:
        for i, row in enumerate(revisions.itertuples(index=False), start=2):
            ws.cell(i, 1, row.indicator)
            ws.cell(i, 2, row.name_ja)
            ws.cell(i, 3, row.region_zh)
            ws.cell(i, 4, row.period)
            ws.cell(i, 5, float(row.old_value))
            ws.cell(i, 6, float(row.new_value))
            ws.cell(i, 7, float(row.diff))
            ws.cell(i, 8, float(row.pct_change)).number_format = PCT_FMT
    widths = [14, 26, 12, 10, 14, 14, 12, 12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
